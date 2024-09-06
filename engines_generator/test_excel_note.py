import unittest
import os
import openpyxl
from datetime import datetime
from excel_note_engine import save_excel_note  # Adjust the import as needed


class TestSaveExcelNote(unittest.TestCase):

    def setUp(self):
        # Define a temporary Excel file for testing
        self.test_file = "test_notes.xlsx"
        global excel_note_file
        excel_note_file = self.test_file

        # Create the Excel file if it doesn't exist
        if not os.path.exists(self.test_file):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Notes"
            wb.save(self.test_file)

    def tearDown(self):
        # Remove the temporary Excel file after each test
        if os.path.exists(self.test_file):
            os.remove(self.test_file)

    def test_save_note(self):
        # Test saving a note
        note = "This is a test note."
        result = save_excel_note(note, self.test_file)

        # Check if the function returns "note saved"
        self.assertEqual(result, "note saved")

        # Verify the note is saved in the Excel file
        wb = openpyxl.load_workbook(self.test_file)
        ws = wb["Notes"]

        # Check the last row for the note and timestamp
        last_row = ws.max_row
        timestamp, saved_note = (
            ws.cell(row=last_row, column=1).value,
            ws.cell(row=last_row, column=2).value,
        )
        print("last cell col 1")
        print(ws.cell(row=last_row, column=1).value)
        print("last cell col 2")
        print(ws.cell(row=last_row, column=2).value)

        # Verify the note and timestamp
        self.assertEqual(saved_note, note)
        self.assertTrue(datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S"))


if __name__ == "__main__":
    unittest.main()
