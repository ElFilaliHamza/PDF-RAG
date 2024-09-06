from datetime import datetime
import logging
import os

from llama_index.core.tools import FunctionTool
import openpyxl


logger = logging.getLogger()
excel_note_file = "notes.xlsx"  # name of your excel file


def save_excel_note(note, excel_note_file=excel_note_file):
    """
    Save a note to an Excel file.

    Args:
        note (str): The note to be saved.

    Returns:
        str: "note saved" if the note is saved successfully, otherwise False.
    """
    

    try:
        # Make sure the excel file exists
        if not os.path.exists(excel_note_file):
            # Create a new Excel file if it does not exist
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Notes"
            wb.save(excel_note_file)

        # Load the workbook and worksheet
        wb = openpyxl.load_workbook(excel_note_file)
        ws = wb["Notes"]

        # Get the current date and time
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Create a new row with the timestamp and note
        new_row = [timestamp, note]
        row_num = ws.max_row + 1
        # Add the note to the Excel file
        for i, cell_value in enumerate(new_row, start=1):
            # Use max_row to find the next empty row
            ws.cell(row=row_num, column=i).value = cell_value

        # Save the changes
        wb.save(excel_note_file)

        return "note saved"

    except Exception as e:
        # Return False if any error occurs
        return False


excel_note_engine = FunctionTool.from_defaults(
    fn=save_excel_note,
    name="excel_note_saver",
    description="""Simplify note-taking with this powerful tool. Create, save
    and organize your notes in a structured format, then easily export them to
    files for later reference. Ideal for personal or professional use.""",
)
