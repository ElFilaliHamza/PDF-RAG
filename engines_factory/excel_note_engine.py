from datetime import datetime
import logging
import os

from llama_index.core.tools import FunctionTool
import openpyxl


logger = logging.getLogger()
excel_note_file = os.path.join("data", "notes.xlsx")

def save_excel_note(note, excel_note_file=excel_note_file):
    print("Saving note to Excel:", note)
    try:
        # Check if the input is a dictionary and extract the content
        if isinstance(note, dict):
            note_title = note.get("title", "")
            note_content = note.get("content", "")
            note_text = f"{note_title}: {note_content}" if note_title else note_content
        else:
            note_text = note

        # Ensure the directory exists
        os.makedirs(os.path.dirname(excel_note_file), exist_ok=True)

        # Make sure the Excel file exists
        if not os.path.exists(excel_note_file):
            # Create a new Excel file if it does not exist
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Notes"
            wb.save(excel_note_file)

        # Load the workbook and worksheet
        wb = openpyxl.load_workbook(excel_note_file)
        ws = wb["Notes"]

        # Get the current date and time
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Create a new row with the timestamp and note
        new_row = [timestamp, note_text]
        row_num = ws.max_row + 1

        # Add the note to the Excel file
        for i, cell_value in enumerate(new_row, start=1):
            ws.cell(row=row_num, column=i).value = cell_value

        # Save the changes
        wb.save(excel_note_file)

        print("Successfully saved the note to", excel_note_file)
        return "note saved"

    except Exception as e:
        error_message = f"An error occurred while saving the note: {str(e)}"
        print(error_message)
        return error_message


excel_note_engine = FunctionTool.from_defaults(
    fn=save_excel_note,
    name="excel_note_saver",
    description="""This tool saves a note when the user asks for saving a note. 
    The user can save any content they want. 
    The tool is designed to be used in the LLM context.""",
)
